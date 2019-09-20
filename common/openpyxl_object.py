"""
============================
Author:赵健
Date:2019-08-22
Time:22:14
E-mail:948883947@qq.com
File:openpyxl_object.py
============================

"""
import openpyxl



class CaseObject(object):
    '''测试用例对象类'''

    def __init__(self, zip_object):  # zip_object 传入的用例对象列表
        for i in zip_object:
            setattr(self, i[0], i[1])


class ReadExcelData(object):
    """定义读取表格数据类"""

    def __init__(self, excelname, sheetname):
        """
        初始化
        :param excelname: 表格名
        :param sheetname: 表单名
        """
        self.excelname = excelname
        self.sheetname = sheetname

    def open(self):
        '''打开表格'''
        self.workbook = openpyxl.load_workbook(self.excelname)  # 创建工作簿对象
        self.sheet = self.workbook[self.sheetname]  # 创建表单对象

    def read(self):
        '''读取数据'''
        object_list = []  # 定义一个存储用例对象的列表
        self.open()  # 打开表格
        rows_list = list(self.sheet.rows)  # 获取表格的所有行的对象，转化为列表
        title = [row.value for row in rows_list[0]]  # 获取表格title存储为列表
        for row in rows_list[1:]:  # 遍历除第一行的其他行
            data = [r.value for r in row]  # 将其他行数据存储到data列表当中
            zip_object = zip(title, data)  # 将title和data打包
            case_object = CaseObject(zip_object)  # 创建用例对象
            object_list.append(case_object)  # 将用例对象添加到用例对象列表
        return object_list

    def write(self, row, column, value):
        self.open()  # 打开表格
        self.sheet.cell(row=row, column=column, value=value)  # 写入内容
        self.workbook.save(self.excelname)  # 保存


if __name__ == "__main__":
    workbook = openpyxl.load_workbook('testcase.xlsx')
    # print(workbook.sheetnames)
    sheet = workbook['Case']
    print(list(sheet.values))